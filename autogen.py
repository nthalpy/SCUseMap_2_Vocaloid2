# -*- coding: utf-8 -*-
import openpyxl
import datetime

class data:
    displayTitle = None
    titleAnswers = None
    composerName = None
    composerAnswers = None

    releaseDate = None
    singers = None

    timeInMs = None
    
    pass

class lyric:
    timeInMs = None
    lyric = None
    
    pass

endingTimeInMs = None

dataList = []
lyricList = []

wb = openpyxl.load_workbook('data.xlsx')

sheet_song = wb['시트2']
isFirst = True
for row in sheet_song.iter_rows():
    if isFirst:
        isFirst = False
        continue

    d = data()

    d.displayTitle = row[0].value
    if row[1].value != None:
        d.displayTitle = f'{d.displayTitle} ({row[1].value})'

    d.composerName = row[4].value
    if row[5].value != None:
        d.composerName = f'{d.composerName} ({row[5].value})'

    d.titleAnswers = row[2].value.split(',')
    d.composerAnswers = row[6].value.split(',')
    
    d.releaseDate = row[8].value.strftime('%Y년 %m월 %d일')
    d.singers = row[10].value.split(',')

    dataList.append(d)

sheet_lyric = wb['시트3']
isFirst = True
for row in sheet_lyric.iter_rows():
    if isFirst:
        isFirst = False
        continue

    l = lyric()

    l.timeInMs = int(1000 * float(row[0].value))
    l.lyric = f'{row[1].value}   {row[2].value}'

    if row[3].value != None:
        dataIndex = int(row[3].value)
        
        if dataIndex != -1:
            dataList[dataIndex - 1].timeInMs = l.timeInMs
            lyricList.append(l)
        else:
            endingTimeInMs = l.timeInMs
    else:
        lyricList.append(l)           

# 2부터 시작
chatEventStrs = [None, None]

for d in dataList:
    for i in range(len(d.titleAnswers)):
        s = d.titleAnswers[i]
        if not (s in chatEventStrs):
            chatEventStrs.append(s)

        d.titleAnswers[i] = str(chatEventStrs.index(s))
        pass

    for i in range(len(d.composerAnswers)):
        s = d.composerAnswers[i]
        if not (s in chatEventStrs):
            chatEventStrs.append(s)

        d.composerAnswers[i] = str(chatEventStrs.index(s))
        pass
    
for i in range(2, len(chatEventStrs)):
    chatEventStrs[i] = chatEventStrs[i].lower().replace(' ', '', 999)

with open('copy_to_chatEvent.txt', 'w+', encoding='utf8') as copy_to_chatEvent:
    copy_to_chatEvent.write('[chatEvent]\n')
    copy_to_chatEvent.write('__Addr__: 0x58D900\n')
    copy_to_chatEvent.write('__ptrAddr__: 0x58D904\n')
    copy_to_chatEvent.write('__patternAddr__: 0x58D908\n')
    copy_to_chatEvent.write('__lenAddr__: 0x58D90C\n')
    copy_to_chatEvent.write('\n')

    for i in range(2, len(chatEventStrs)):
        copy_to_chatEvent.write(f'{chatEventStrs[i]}: {i}\n')

with open('Source/answer_autogen.eps', 'w+', encoding='utf8') as answer_autogen:
    answer_autogen.write('import answer;\n')
    
    answer_autogen.write('function load() {\n')

    answer_autogen.write(f'answer.table = EUDArray({len(dataList)});\n')
    answer_autogen.write(f'answer.tableSize = {len(dataList)};\n')
    answer_autogen.write('\n')

    for idx in range(len(dataList)):
        d = dataList[idx]

        answer_autogen.write(
            f'answer.__setData(' +
            f'{idx}, Db("{d.displayTitle}"), [{",".join(d.titleAnswers)}], Db("{d.composerName}"), [{",".join(d.composerAnswers)}], {d.timeInMs}, Db("{d.releaseDate}"));' + 
            '\n')

    answer_autogen.write('}')

with open('Source/lyric_autogen.eps', 'w+', encoding='utf8') as lyric_autogen:
    lyric_autogen.write('import lyric;\n')

    lyric_autogen.write('function load() {\n')

    lyric_autogen.write(f'lyric.table = EUDArray({len(lyricList)});\n')
    lyric_autogen.write(f'lyric.tableSize = {len(lyricList)};\n')
    lyric_autogen.write(f'lyric.endingTimeInMs = {endingTimeInMs};\n')
    lyric_autogen.write('\n')

    for idx in range(len(lyricList)):
        l = lyricList[idx]

        replaced = l.lyric.replace('"', '\\"', 999)
        lyric_autogen.write(f'lyric.__setData({idx}, {l.timeInMs}, Db("{replaced}"));\n')

    lyric_autogen.write('}')


